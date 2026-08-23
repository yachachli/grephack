#!/usr/bin/env python3
"""Normalize the vineyard workbooks into stable Engineer 1 JSON contracts."""

from __future__ import annotations

import hashlib
import json
import math
import re
import warnings
from collections import Counter, defaultdict
from dataclasses import dataclass
from datetime import date, datetime
from pathlib import Path
from statistics import mean
from typing import Any, Iterable

from openpyxl import load_workbook

warnings.filterwarnings(
    "ignore",
    category=UserWarning,
    module=r"openpyxl\.worksheet\._reader",
)


ROOT = Path(__file__).resolve().parents[1]
DATA = ROOT / "data"
OUTPUT = ROOT / "engineer_1" / "generated"
DEFAULT_PAYLOAD_TONS = 23.5


def clean(value: Any) -> str | None:
    if value is None:
        return None
    text = re.sub(r"\s+", " ", str(value)).strip()
    return text or None


def normalized_name(value: Any) -> str | None:
    text = clean(value)
    if not text:
        return None
    return re.sub(r"[^A-Z0-9]+", "", text.upper()) or None


def header_key(value: Any) -> str | None:
    """Normalize decorative spreadsheet headers such as '\nVINEYARD NAME'."""
    return normalized_name(value)


def code(value: Any) -> str | None:
    text = clean(value)
    return re.sub(r"\s+", "", text.upper()) if text else None


def number(value: Any) -> float | None:
    if isinstance(value, bool) or value is None:
        return None
    if isinstance(value, (int, float)) and math.isfinite(float(value)):
        return float(value)
    if isinstance(value, str):
        match = re.fullmatch(r"\s*-?\d+(?:\.\d+)?\s*", value)
        if match:
            return float(value)
    return None


def iso_date(value: Any) -> str | None:
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    return None


def stable_id(prefix: str, *parts: Any) -> str:
    raw = "|".join("" if part is None else str(part) for part in parts)
    digest = hashlib.sha1(raw.encode("utf-8")).hexdigest()[:12]
    return f"{prefix}:{digest}"


def write_json(name: str, payload: Any) -> None:
    OUTPUT.mkdir(parents=True, exist_ok=True)
    (OUTPUT / name).write_text(
        json.dumps(payload, indent=2, sort_keys=True, ensure_ascii=False) + "\n",
        encoding="utf-8",
    )


@dataclass
class BlockIndexes:
    by_subblock: dict[str, dict[str, Any]]
    by_parent: dict[str, list[dict[str, Any]]]


def read_blocks() -> tuple[list[dict[str, Any]], list[dict[str, Any]], BlockIndexes]:
    workbook = load_workbook(
        DATA / "2026 HARVEST STATUS.xlsx", read_only=True, data_only=True
    )
    sheet = workbook["Source List"]
    blocks: list[dict[str, Any]] = []
    seen_subblocks: set[str] = set()

    for row_number, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        parent_code = code(row[5] if len(row) > 5 else None)
        subblock_code = code(row[6] if len(row) > 6 else None)
        if not parent_code or not subblock_code or subblock_code in seen_subblocks:
            continue
        seen_subblocks.add(subblock_code)
        variety_match = re.search(r"([A-Z]{2})$", parent_code)
        blocks.append(
            {
                "block_id": f"block:{parent_code}",
                "subblock_id": f"subblock:{subblock_code}",
                "parent_code": parent_code,
                "subblock_code": subblock_code,
                "region": clean(row[7] if len(row) > 7 else None),
                "variety_code": variety_match.group(1) if variety_match else None,
                "vines_per_acre": number(row[8] if len(row) > 8 else None),
                "total_vines": number(row[9] if len(row) > 9 else None),
                "total_acres": number(row[10] if len(row) > 10 else None),
                "active_2026": None,
                "source": {
                    "workbook": "2026 HARVEST STATUS.xlsx",
                    "sheet": "Source List",
                    "row": row_number,
                },
            }
        )

    by_parent: dict[str, list[dict[str, Any]]] = defaultdict(list)
    by_subblock: dict[str, dict[str, Any]] = {}
    for block in blocks:
        by_parent[block["parent_code"]].append(block)
        by_subblock[block["subblock_code"]] = block

    parent_blocks: list[dict[str, Any]] = []
    for parent_code, children in sorted(by_parent.items()):
        regions = [child["region"] for child in children if child["region"]]
        varieties = [child["variety_code"] for child in children if child["variety_code"]]
        parent_blocks.append(
            {
                "block_id": f"block:{parent_code}",
                "parent_code": parent_code,
                "region": Counter(regions).most_common(1)[0][0] if regions else None,
                "variety_code": Counter(varieties).most_common(1)[0][0] if varieties else None,
                "total_acres": round(
                    sum(child["total_acres"] or 0 for child in children), 4
                ),
                "subblock_count": len(children),
                "active_2026": None,
            }
        )
    workbook.close()
    return blocks, parent_blocks, BlockIndexes(by_subblock, dict(by_parent))


def read_harvest_events(indexes: BlockIndexes) -> list[dict[str, Any]]:
    workbook = load_workbook(
        DATA / "2026 HARVEST STATUS.xlsx", read_only=True, data_only=True
    )
    sheet = workbook["Row Counts"]
    events: list[dict[str, Any]] = []
    for row_number, row in enumerate(sheet.iter_rows(min_row=3, values_only=True), start=3):
        observed_at = iso_date(row[0] if row else None)
        source_code = code(row[1] if len(row) > 1 else None)
        if not observed_at or not source_code:
            continue
        matched = indexes.by_subblock.get(source_code)
        actual_loads = number(row[10] if len(row) > 10 else None)
        actual_tons = number(row[12] if len(row) > 12 else None)
        payload = (
            round(actual_tons / actual_loads, 4)
            if actual_tons is not None and actual_loads and actual_loads > 0
            else None
        )
        events.append(
            {
                "event_id": stable_id("event", observed_at, source_code, row_number),
                "observed_at": observed_at,
                "source_block_code": source_code,
                "block_id": matched["block_id"] if matched else None,
                "subblock_id": matched["subblock_id"] if matched else None,
                "vines_per_acre": number(row[2] if len(row) > 2 else None),
                "total_vines": number(row[3] if len(row) > 3 else None),
                "vines_picked": number(row[4] if len(row) > 4 else None),
                "vines_remaining": number(row[5] if len(row) > 5 else None),
                "total_acres": number(row[6] if len(row) > 6 else None),
                "acres_picked": number(row[7] if len(row) > 7 else None),
                "acres_remaining": number(row[8] if len(row) > 8 else None),
                "acres_picked_since_last_count": number(row[9] if len(row) > 9 else None),
                "actual_loads": actual_loads,
                "winery_raw": clean(row[11] if len(row) > 11 else None),
                "actual_tons": actual_tons,
                "observed_tpa": number(row[13] if len(row) > 13 else None),
                "workbook_loads_to_finish": number(row[14] if len(row) > 14 else None),
                "payload_tons": payload,
                "reported_by_raw": clean(row[15] if len(row) > 15 else None),
                "rows_raw": clean(row[16] if len(row) > 16 else None),
                "notes": clean(row[17] if len(row) > 17 else None),
                "source": {
                    "workbook": "2026 HARVEST STATUS.xlsx",
                    "sheet": "Row Counts",
                    "row": row_number,
                },
            }
        )
    workbook.close()
    return events


def derive_block_status(events: list[dict[str, Any]]) -> list[dict[str, Any]]:
    grouped: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for event in events:
        key = event["subblock_id"] or f"source:{event['source_block_code']}"
        grouped[key].append(event)

    statuses: list[dict[str, Any]] = []
    for group_events in grouped.values():
        ordered = sorted(
            group_events,
            key=lambda item: (item["observed_at"], item["source"]["row"]),
        )
        latest = ordered[-1]
        usable = [
            event
            for event in reversed(ordered)
            if (event["acres_picked_since_last_count"] or 0) > 0
            and (event["actual_tons"] or 0) > 0
        ][:3]
        acreage = sum(event["acres_picked_since_last_count"] or 0 for event in usable)
        tons = sum(event["actual_tons"] or 0 for event in usable)
        rolling_tpa = tons / acreage if acreage > 0 else None
        acres_remaining = latest["acres_remaining"]
        estimated_tons = (
            acres_remaining * rolling_tpa
            if acres_remaining is not None and rolling_tpa is not None
            else None
        )
        estimated_loads = (
            estimated_tons / DEFAULT_PAYLOAD_TONS
            if estimated_tons is not None
            else None
        )
        acres_picked = latest["acres_picked"]
        if acres_remaining is not None and acres_remaining <= 0.1:
            status = "harvested"
        elif acres_picked is not None and acres_picked > 0:
            status = "in_progress"
        else:
            status = "pending"
        statuses.append(
            {
                "block_id": latest["block_id"],
                "subblock_id": latest["subblock_id"],
                "source_block_code": latest["source_block_code"],
                "as_of_date": latest["observed_at"],
                "status": status,
                "total_acres": latest["total_acres"],
                "acres_picked": acres_picked,
                "acres_remaining": acres_remaining,
                "rolling_tpa": round(rolling_tpa, 4) if rolling_tpa is not None else None,
                "estimated_tons_remaining": round(estimated_tons, 2)
                if estimated_tons is not None
                else None,
                "estimated_loads_remaining": round(estimated_loads, 2)
                if estimated_loads is not None
                else None,
                "required_trucks_one_trip": math.ceil(estimated_loads)
                if estimated_loads is not None
                else None,
                "workbook_loads_to_finish": latest["workbook_loads_to_finish"],
                "payload_assumption_tons": DEFAULT_PAYLOAD_TONS,
                "observations_used_for_rolling_tpa": len(usable),
            }
        )
    return sorted(statuses, key=lambda item: item["source_block_code"])


def find_header_row(sheet: Any, required: str, max_rows: int = 10) -> tuple[int, list[Any]]:
    required_key = header_key(required)
    accepted = {required_key, f"{required_key}NAME"}
    for row_number, row in enumerate(
        sheet.iter_rows(min_row=1, max_row=max_rows, values_only=True), start=1
    ):
        values = [header_key(value) for value in row]
        if any(value in accepted for value in values):
            return row_number, list(row)
    raise ValueError(f"Could not find {required!r} header in {sheet.title}")


def parse_sample_date(value: Any, season: int) -> str | None:
    direct = iso_date(value)
    if direct:
        return direct
    text = clean(value)
    if not text:
        return None
    text = re.sub(r"^[A-Za-z]{3,9}[, ]+", "", text)
    for fmt in ("%d-%b-%Y", "%d-%b", "%b-%d", "%m/%d/%Y", "%m/%d"):
        try:
            parsed = datetime.strptime(text, fmt)
            year = parsed.year if "%Y" in fmt else season
            return date(year, parsed.month, parsed.day).isoformat()
        except ValueError:
            pass
    return None


def parse_brix(value: Any) -> tuple[float | None, str]:
    if isinstance(value, bool) or value is None:
        return None, "missing"
    if isinstance(value, (int, float)):
        result = float(value)
        return (result, "numeric") if 5 <= result <= 40 else (None, "out_of_range")
    text = clean(value) or ""
    candidates = [float(item) for item in re.findall(r"(?<!\d)(\d{1,2}(?:\.\d+)?)", text)]
    plausible = [item for item in candidates if 5 <= item <= 40]
    if not plausible:
        return None, "unparsed"
    return mean(plausible), "annotated" if len(plausible) > 1 or not re.fullmatch(r"\d+(?:\.\d+)?", text) else "numeric_text"


def read_brix(indexes: BlockIndexes) -> tuple[list[dict[str, Any]], dict[str, set[str]]]:
    sources = [
        (2022, "2022 SUGARS.xlsx", "Sheet1"),
        (2023, "2023 SUGARS (1).xlsx", "All Brix"),
        (2024, "2024 SUGARS.xlsx", "All Brix"),
        (2026, "2026 SUGARS.xlsx", "ALL Sugars"),
    ]
    observations: list[dict[str, Any]] = []
    name_to_codes: dict[str, set[str]] = defaultdict(set)

    aliases = {
        "CODE": {"CODE", "BLOCK CODE"},
        "VINEYARD": {"VINEYARD", "VINEYARD NAME"},
        "VARIETY": {"VARIETY", "VAR"},
        "WINERY": {"WINERY", "CLIENT"},
    }

    for season, filename, sheet_name in sources:
        workbook = load_workbook(DATA / filename, read_only=True, data_only=True)
        sheet = workbook[sheet_name]
        header_row, header = find_header_row(sheet, "VINEYARD")
        header_names = [header_key(value) for value in header]

        def locate(field: str) -> int | None:
            for position, name in enumerate(header_names):
                if name in {header_key(alias) for alias in aliases[field]}:
                    return position
            return None

        code_col = locate("CODE")
        vineyard_col = locate("VINEYARD")
        variety_col = locate("VARIETY")
        winery_col = locate("WINERY")
        date_columns = {
            position: parsed
            for position, value in enumerate(header)
            if (parsed := parse_sample_date(value, season)) is not None
        }

        for row_number, row in enumerate(
            sheet.iter_rows(min_row=header_row + 1, values_only=True),
            start=header_row + 1,
        ):
            vineyard = clean(row[vineyard_col]) if vineyard_col is not None and len(row) > vineyard_col else None
            source_code = code(row[code_col]) if code_col is not None and len(row) > code_col else None
            if not vineyard:
                continue
            vineyard_key = normalized_name(vineyard)
            if source_code and vineyard_key:
                name_to_codes[vineyard_key].add(source_code)
            block_id = f"block:{source_code}" if source_code in indexes.by_parent else None
            for column, sampled_at in date_columns.items():
                if column >= len(row):
                    continue
                parsed_brix, parse_kind = parse_brix(row[column])
                if parsed_brix is None:
                    continue
                raw = row[column]
                raw_text = clean(raw) or ""
                provenance = "external_or_annotated" if "*" in raw_text else "field_or_unspecified"
                observations.append(
                    {
                        "observation_id": stable_id(
                            "brix",
                            filename,
                            row_number,
                            season,
                            source_code,
                            vineyard,
                            sampled_at,
                            column,
                        ),
                        "season": season,
                        "sampled_at": sampled_at,
                        "source_code": source_code,
                        "block_id": block_id,
                        "vineyard_name": vineyard,
                        "vineyard_key": vineyard_key,
                        "variety": clean(row[variety_col])
                        if variety_col is not None and len(row) > variety_col
                        else None,
                        "winery_raw": clean(row[winery_col])
                        if winery_col is not None and len(row) > winery_col
                        else None,
                        "brix": round(parsed_brix, 3),
                        "raw_value": raw,
                        "parse_kind": parse_kind,
                        "provenance": provenance,
                        "source": {
                            "workbook": filename,
                            "sheet": sheet_name,
                            "row": row_number,
                            "column": column + 1,
                        },
                    }
                )
        workbook.close()
    return observations, name_to_codes


def read_crop_estimates(indexes: BlockIndexes) -> list[dict[str, Any]]:
    workbook = load_workbook(
        DATA / "2026 Crop Estimation.xlsx", read_only=True, data_only=True
    )
    sheet = workbook["2026 Crop Estimation"]
    estimates: list[dict[str, Any]] = []
    for row_number, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        source_code = code(row[1] if len(row) > 1 else None)
        block_name = clean(row[2] if len(row) > 2 else None)
        if not source_code or not block_name:
            continue
        samples = [
            value
            for value in (number(item) for item in row[34:90])
            if value is not None
        ]
        net_acres = number(row[8] if len(row) > 8 else None)
        estimated_tpa = number(row[29] if len(row) > 29 else None)
        estimated_total = (
            net_acres * estimated_tpa
            if net_acres is not None and estimated_tpa is not None
            else None
        )
        variety = clean(row[4] if len(row) > 4 else None)
        series_key = "|".join(
            [source_code, normalized_name(block_name) or "", normalized_name(variety) or ""]
        )
        estimates.append(
            {
                "observation_id": stable_id("crop", series_key, row_number),
                "series_key": series_key,
                "block_id": f"block:{source_code}" if source_code in indexes.by_parent else None,
                "source_code": source_code,
                "source_block_name": block_name,
                "region": clean(row[0] if row else None),
                "variety": variety,
                "clone": clean(row[5] if len(row) > 5 else None),
                "rootstock": clean(row[6] if len(row) > 6 else None),
                "development_status": clean(row[3] if len(row) > 3 else None),
                "vines_per_acre": number(row[7] if len(row) > 7 else None),
                "net_acres": net_acres,
                "adjusted_vines": number(row[11] if len(row) > 11 else None),
                "observed_at": iso_date(row[26] if len(row) > 26 else None),
                "workbook_average_clusters": number(row[27] if len(row) > 27 else None),
                "estimated_lbs_per_vine": number(row[28] if len(row) > 28 else None),
                "estimated_tpa": estimated_tpa,
                "estimated_total_tons": round(estimated_total, 3)
                if estimated_total is not None
                else None,
                "actual_tpa_raw": number(row[30] if len(row) > 30 else None),
                "sample_n": len(samples),
                "sample_mean_all_vines": round(mean(samples), 4) if samples else None,
                "selected": False,
                "notes": clean(row[32] if len(row) > 32 else None),
                "counter_raw": clean(row[33] if len(row) > 33 else None),
                "source": {
                    "workbook": "2026 Crop Estimation.xlsx",
                    "sheet": "2026 Crop Estimation",
                    "row": row_number,
                },
            }
        )

    grouped: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for estimate in estimates:
        grouped[estimate["series_key"]].append(estimate)
    for group in grouped.values():
        dated = [item for item in group if item["observed_at"]]
        if dated:
            selected = max(
                dated, key=lambda item: (item["observed_at"], item["source"]["row"])
            )
            selected["selected"] = True
    workbook.close()
    return estimates


SCHEDULE_HEADING = re.compile(
    r"^(?:MONDAY|TUESDAY|WEDNESDAY|THURSDAY|FRIDAY|SATURDAY|SUNDAY),\s+"
    r"([A-Z]+)\s+(\d{1,2})$",
    re.IGNORECASE,
)


def schedule_date(value: Any) -> str | None:
    text = clean(value)
    if not text:
        return None
    match = SCHEDULE_HEADING.match(text)
    if not match:
        return None
    try:
        parsed = datetime.strptime(f"{match.group(1)} {match.group(2)} 2026", "%B %d %Y")
        return parsed.date().isoformat()
    except ValueError:
        return None


def schedule_snapshot_date(sheet: Any) -> str | None:
    for cell in ("A3", "A6"):
        value = sheet[cell].value
        parsed = iso_date(value)
        if parsed:
            return parsed
    return None


def read_schedule(
    indexes: BlockIndexes, name_to_codes: dict[str, set[str]]
) -> list[dict[str, Any]]:
    workbook = load_workbook(
        DATA / "2026 HARVEST SCHEDULE.xlsx", read_only=True, data_only=True
    )
    plans: list[dict[str, Any]] = []
    skip = {
        "VINEYARD",
        "UPCOMING DELIVERIES",
        "HARVEST SCHEDULE 2026",
    }
    for sheet in workbook.worksheets:
        if sheet.title == "Source List" or sheet.sheet_state != "visible":
            continue
        snapshot = schedule_snapshot_date(sheet)
        current_plan_date: str | None = None
        for row_number, row in enumerate(
            sheet.iter_rows(min_row=1, max_row=min(sheet.max_row, 250), values_only=True),
            start=1,
        ):
            first = clean(row[0] if row else None)
            heading_date = schedule_date(first)
            if heading_date:
                current_plan_date = heading_date
                continue
            if not first or first.upper() in skip or not current_plan_date:
                continue
            if first.upper().startswith(("DELIVERY STATUS", "PRINTED:", "1 MACHINE")):
                continue
            operational = [row[i] if len(row) > i else None for i in range(1, 9)]
            if not any(value not in (None, "") for value in operational):
                continue
            vineyard_key = normalized_name(first)
            candidates = name_to_codes.get(vineyard_key or "", set())
            matched_code = next(iter(candidates)) if len(candidates) == 1 else None
            block_id = (
                f"block:{matched_code}"
                if matched_code and matched_code in indexes.by_parent
                else None
            )
            raw_loads = row[4] if len(row) > 4 else None
            numeric_loads = number(raw_loads)
            plans.append(
                {
                    "plan_id": stable_id(
                        "plan", sheet.title, row_number, current_plan_date, first
                    ),
                    "snapshot_date": snapshot,
                    "snapshot_sheet": sheet.title,
                    "plan_date": current_plan_date,
                    "vineyard_name": first,
                    "vineyard_key": vineyard_key,
                    "source_code": matched_code,
                    "block_id": block_id,
                    "match_method": "exact_sugar_name"
                    if block_id
                    else "unresolved",
                    "harvester_or_start_raw": clean(row[1] if len(row) > 1 else None),
                    "stop_brix_raw": clean(row[2] if len(row) > 2 else None),
                    "notes": clean(row[3] if len(row) > 3 else None),
                    "scheduled_loads_raw": raw_loads,
                    "scheduled_loads": numeric_loads,
                    "flag_raw": clean(row[5] if len(row) > 5 else None),
                    "winery_raw": clean(row[6] if len(row) > 6 else None),
                    "delivery_times_raw": clean(row[7] if len(row) > 7 else None),
                    "confirmed_raw": clean(row[8] if len(row) > 8 else None),
                    "source": {
                        "workbook": "2026 HARVEST SCHEDULE.xlsx",
                        "sheet": sheet.title,
                        "row": row_number,
                    },
                }
            )
    workbook.close()
    return plans


def quality_report(
    blocks: list[dict[str, Any]],
    parent_blocks: list[dict[str, Any]],
    events: list[dict[str, Any]],
    statuses: list[dict[str, Any]],
    brix: list[dict[str, Any]],
    crop: list[dict[str, Any]],
    schedule: list[dict[str, Any]],
    backend_candidates: list[dict[str, Any]],
) -> dict[str, Any]:
    return {
        "generated_at": datetime.now().astimezone().isoformat(timespec="seconds"),
        "counts": {
            "parent_blocks": len(parent_blocks),
            "subblocks": len(blocks),
            "harvest_events": len(events),
            "block_status_records": len(statuses),
            "brix_observations": len(brix),
            "crop_estimate_observations": len(crop),
            "selected_crop_estimates": sum(bool(item["selected"]) for item in crop),
            "schedule_rows": len(schedule),
            "backend_block_candidates": len(backend_candidates),
            "backend_api_ready_candidates": sum(
                item["apiReady"] for item in backend_candidates
            ),
        },
        "coverage": {
            "managed_acres_from_status_master": round(
                sum(item["total_acres"] or 0 for item in blocks), 2
            ),
            "unmatched_harvest_events": sum(item["block_id"] is None for item in events),
            "unmatched_brix_observations": sum(item["block_id"] is None for item in brix),
            "unmatched_crop_observations": sum(item["block_id"] is None for item in crop),
            "unmatched_schedule_rows": sum(item["block_id"] is None for item in schedule),
            "schedule_rows_with_numeric_loads": sum(
                item["scheduled_loads"] is not None for item in schedule
            ),
        },
        "assumptions": [
            "Status Source List is the canonical acreage/block master.",
            "Status latest observation is the current operational state.",
            "Each truck is assumed to make one trip per night and carry 23.5 tons; required loads therefore equal required trucks for the MVP.",
            "Variety suffixes derived from Status parent codes are provisional.",
            "Crop selected estimate is the latest dated row per code + normalized block name + variety.",
            "Schedule joins require an exact, unambiguous normalized name from the Sugar workbook.",
            "2022 Sugar observations remain parent-unmapped because that workbook has no codes.",
        ],
        "known_source_risks": [
            "Harvest Summary covers a smaller implicit scope than the Status block master.",
            "Schedule load cells mix numeric loads with tons, bins, and question marks.",
            "Source formulas mix 23, 23.5, and 24 tons per load; the operator-confirmed practical payload is 23.5 tons despite four nominal six-ton compartments.",
            "Crop code is not unique at observation grain and some rows mix units.",
            "Sugar field and lab methods are not interchangeable without calibration.",
            "Historical Brix backtests must avoid LAST BRIX, LAST SAMPLE DATE, and DONE fields to prevent leakage.",
        ],
    }


def build_backend_candidates(
    parent_blocks: list[dict[str, Any]],
    brix: list[dict[str, Any]],
    crop: list[dict[str, Any]],
    schedule: list[dict[str, Any]],
) -> list[dict[str, Any]]:
    names_by_code: dict[str, Counter[str]] = defaultdict(Counter)
    for observation in brix:
        if observation["season"] == 2026 and observation["source_code"]:
            names_by_code[observation["source_code"]][observation["vineyard_name"]] += 1

    selected_crop: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for estimate in crop:
        if estimate["selected"] and estimate["estimated_total_tons"] is not None:
            selected_crop[estimate["source_code"]].append(estimate)

    snapshot_dates = [item["snapshot_date"] for item in schedule if item["snapshot_date"]]
    latest_snapshot = max(snapshot_dates) if snapshot_dates else None
    plan_dates: dict[str, list[str]] = defaultdict(list)
    for item in schedule:
        if (
            item["block_id"]
            and item["snapshot_date"] == latest_snapshot
            and item["plan_date"]
        ):
            plan_dates[item["block_id"]].append(item["plan_date"])

    candidates: list[dict[str, Any]] = []
    for parent in parent_blocks:
        parent_code = parent["parent_code"]
        name_counts = names_by_code.get(parent_code)
        vineyard_name = (
            sorted(name_counts.items(), key=lambda item: (-item[1], len(item[0]), item[0]))[0][0]
            if name_counts
            else parent_code
        )
        crop_rows = selected_crop.get(parent_code, [])
        estimated_tons = (
            crop_rows[0]["estimated_total_tons"] if len(crop_rows) == 1 else None
        )
        dates = sorted(set(plan_dates.get(parent["block_id"], [])))
        harvest_start = dates[0] if dates else None
        harvest_end = dates[-1] if dates else None
        blockers: list[str] = []
        if not crop_rows:
            blockers.append("missing_selected_crop_estimate")
        elif len(crop_rows) > 1:
            blockers.append("multiple_crop_segments_require_rollup_review")
        if not dates:
            blockers.append("missing_harvest_window_from_forecasting_or_schedule")
        candidates.append(
            {
                "externalId": parent_code,
                "vineyardName": vineyard_name,
                "blockName": parent_code,
                "variety": parent["variety_code"] or "UNKNOWN",
                "region": parent["region"] or "UNKNOWN",
                "acres": round(parent["total_acres"], 2),
                "estimatedTons": estimated_tons,
                "harvestWindowStart": harvest_start,
                "harvestWindowEnd": harvest_end,
                "apiReady": not blockers,
                "blockers": blockers,
            }
        )
    return candidates


def build_review_queue(
    events: list[dict[str, Any]],
    crop: list[dict[str, Any]],
    schedule: list[dict[str, Any]],
) -> dict[str, Any]:
    unmatched_schedule: dict[str, dict[str, Any]] = {}
    for item in schedule:
        if item["block_id"] is not None:
            continue
        key = item["vineyard_key"] or item["vineyard_name"]
        record = unmatched_schedule.setdefault(
            key,
            {
                "vineyard_name": item["vineyard_name"],
                "occurrences": 0,
                "winery_values": set(),
                "plan_dates": set(),
            },
        )
        record["occurrences"] += 1
        if item["winery_raw"]:
            record["winery_values"].add(item["winery_raw"])
        if item["plan_date"]:
            record["plan_dates"].add(item["plan_date"])

    schedule_review = []
    for record in unmatched_schedule.values():
        schedule_review.append(
            {
                **record,
                "winery_values": sorted(record["winery_values"]),
                "plan_dates": sorted(record["plan_dates"]),
            }
        )
    schedule_review.sort(key=lambda item: (-item["occurrences"], item["vineyard_name"]))

    return {
        "unmatched_harvest_event_codes": sorted(
            {item["source_block_code"] for item in events if item["block_id"] is None}
        ),
        "unmatched_crop_codes": sorted(
            {item["source_code"] for item in crop if item["block_id"] is None}
        ),
        "unmatched_schedule_names": schedule_review,
    }


def run() -> dict[str, Any]:
    blocks, parent_blocks, indexes = read_blocks()
    events = read_harvest_events(indexes)
    statuses = derive_block_status(events)
    brix, name_to_codes = read_brix(indexes)
    crop = read_crop_estimates(indexes)
    for estimate in crop:
        alias_key = normalized_name(estimate["source_block_name"])
        if alias_key:
            name_to_codes[alias_key].add(estimate["source_code"])
    schedule = read_schedule(indexes, name_to_codes)
    backend_candidates = build_backend_candidates(parent_blocks, brix, crop, schedule)
    review_queue = build_review_queue(events, crop, schedule)
    report = quality_report(
        blocks,
        parent_blocks,
        events,
        statuses,
        brix,
        crop,
        schedule,
        backend_candidates,
    )

    write_json("parent_blocks.json", parent_blocks)
    write_json("blocks.json", blocks)
    write_json("harvest_events.json", events)
    write_json("block_status.json", statuses)
    write_json("brix_observations.json", brix)
    write_json("crop_estimates.json", crop)
    write_json("harvest_plan.json", schedule)
    write_json("backend_block_candidates.json", backend_candidates)
    write_json("review_queue.json", review_queue)
    write_json("data_quality.json", report)
    return report


if __name__ == "__main__":
    print(json.dumps(run(), indent=2, sort_keys=True))
